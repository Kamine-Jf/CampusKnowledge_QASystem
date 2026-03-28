# -*- coding: utf-8 -*-
"""Excel 数据导入 MySQL 的批处理脚本。"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

from .mysql_conn import MysqlConnection

try:
    import openpyxl  # type: ignore
except ImportError:
    openpyxl = None

try:
    import xlrd  # type: ignore
except ImportError:
    xlrd = None

# Excel 数据目录常量，固定读取项目结构中的 structured_data 目录。
EXCEL_ROOT = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "data", "structured_data")
EXCEL_ROOT = os.path.abspath(EXCEL_ROOT)

# Excel 表头与数据库字段的映射关系，确保按表头匹配数据而非固定列序。
HEADER_FIELD_MAP = {
    "大类分类": "category",
    "具体事项": "item",
    "详细操作/说明": "operation",
    "时间要求/关键节点": "time_requirement",
    "办理渠道/联系方式": "channel",
    "信息来源/备注": "source_note",
}


def _normalize_header(header_value: Any) -> str:
    """规范化表头名称，消除空格与换行差异。"""
    if header_value is None:
        return ""
    return str(header_value).replace("\r", "\n").replace("\n", "").strip()


def _clean_cell(cell_value: Any) -> str:
    """清洗单元格数据，统一空值与空白字符。"""
    if cell_value is None:
        return ""
    text = str(cell_value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def _collect_excel_files(root_dir: str) -> List[str]:
    """收集指定目录下全部 Excel 文件路径。"""
    if not os.path.isdir(root_dir):
        raise FileNotFoundError(f"未找到 Excel 目录：{root_dir}")
    excel_files: List[str] = []
    for entry in os.listdir(root_dir):
        if entry.lower().endswith((".xlsx", ".xls")):
            excel_files.append(os.path.join(root_dir, entry))
    return excel_files


def _parse_xlsx(file_path: str) -> List[Dict[str, str]]:
    """解析 .xlsx 文件的所有工作表。"""
    if openpyxl is None:
        raise ImportError("未安装 openpyxl 库，无法解析 .xlsx 文件。")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    records: List[Dict[str, str]] = []
    for sheet in workbook.worksheets:
        if sheet.max_row is None or sheet.max_row < 2:
            continue
        headers = [
            _normalize_header(cell.value)
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        header_index_map = _build_header_index_map(headers)
        for row in sheet.iter_rows(min_row=2):
            row_values = [_clean_cell(cell.value) for cell in row]
            record = _build_record_from_row(row_values, header_index_map)
            if record:
                records.append(record)
    return records


def _parse_xls(file_path: str) -> List[Dict[str, str]]:
    """解析 .xls 文件的所有工作表。"""
    if xlrd is None:
        raise ImportError("未安装 xlrd 库，无法解析 .xls 文件。")
    workbook = xlrd.open_workbook(file_path)
    records: List[Dict[str, str]] = []
    for sheet in workbook.sheets():
        if sheet.nrows < 2:
            continue
        headers = [_normalize_header(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        header_index_map = _build_header_index_map(headers)
        for row_idx in range(1, sheet.nrows):
            row_values = [_clean_cell(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
            record = _build_record_from_row(row_values, header_index_map)
            if record:
                records.append(record)
    return records


def _build_header_index_map(headers: Sequence[str]) -> Dict[str, int]:
    """根据表头生成列索引映射，确保列顺序变化时仍能准确匹配。"""
    index_map: Dict[str, int] = {}
    for index, header in enumerate(headers):
        if header in HEADER_FIELD_MAP:
            index_map[HEADER_FIELD_MAP[header]] = index
    missing_headers = set(HEADER_FIELD_MAP.values()) - set(index_map.keys())
    if missing_headers:
        raise ValueError(f"表头缺失，无法匹配列：{missing_headers}")
    return index_map


def _build_record_from_row(row_values: Sequence[str], header_index_map: Dict[str, int]) -> Optional[Dict[str, str]]:
    """基于列索引映射构建单条记录，自动跳过空行。"""
    record: Dict[str, str] = {}
    for field_name, column_index in header_index_map.items():
        if column_index >= len(row_values):
            record[field_name] = ""
        else:
            record[field_name] = row_values[column_index]
    # 若核心字段均为空，则视为无效行直接忽略。
    if not record.get("category") and not record.get("item") and not record.get("operation"):
        return None
    return record


def _deduplicate_records(records: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """按照 category+item 组合去重，防止重复入库。"""
    unique_records: List[Dict[str, str]] = []
    seen_keys: Set[Tuple[str, str]] = set()
    for record in records:
        key = (record.get("category", ""), record.get("item", ""))
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_records.append(record)
    return unique_records


def _read_single_file(file_path: str) -> List[Dict[str, str]]:
    """根据扩展名选择解析器，返回该文件的全部记录。"""
    if file_path.lower().endswith(".xlsx"):
        return _parse_xlsx(file_path)
    if file_path.lower().endswith(".xls"):
        return _parse_xls(file_path)
    raise ValueError(f"不支持的文件类型：{file_path}")


def excel_to_mysql() -> int:
    """主入口函数：读取 Excel、清洗数据并批量入库。

    返回值说明：
        int: 本次成功写入数据库的记录条数，用于后续测试脚本统计。

    异常处理：
        捕获所有解析与入库异常，打印详细错误后返回 0，确保上层流程不中断。
    """
    try:
        excel_files = _collect_excel_files(EXCEL_ROOT)
    except (FileNotFoundError, PermissionError) as directory_error:
        print(f"[目录错误] {directory_error}")
        return 0

    if not excel_files:
        print("[数据提示] structured_data 目录下未发现 Excel 文件。")
        return 0

    all_records: List[Dict[str, str]] = []
    for file_path in excel_files:
        try:
            records = _read_single_file(file_path)
            if records:
                print(f"[解析成功] {os.path.basename(file_path)} 共读取 {len(records)} 条记录。")
                all_records.extend(records)
            else:
                print(f"[解析提示] {os.path.basename(file_path)} 未读取到有效数据。")
        except Exception as parse_error:  # pylint: disable=broad-except
            print(f"[解析失败] 文件 {file_path} 发生错误：{parse_error}")

    if not all_records:
        print("[数据提示] 未收集到任何可入库的数据。")
        return 0

    unique_records = _deduplicate_records(all_records)
    print(f"[数据预处理] 去重后剩余 {len(unique_records)} 条记录。")

    mysql_client = MysqlConnection()
    connection_tuple = mysql_client.connect_db()
    if connection_tuple is None:
        print("[连接失败] 无法建立数据库连接，终止导入。")
        return 0
    connection, cursor = connection_tuple

    insert_sql = (
        "INSERT INTO campus_struct_data (category, item, operation, time_requirement, channel, source_note) "
        "VALUES (%(category)s, %(item)s, %(operation)s, %(time_requirement)s, %(channel)s, %(source_note)s)"
    )

    try:
        cursor.executemany(insert_sql, unique_records)
        connection.commit()
        print(f"[入库成功] 成功写入 {cursor.rowcount} 条记录。")
        return cursor.rowcount
    except Exception as import_error:  # pylint: disable=broad-except
        connection.rollback()
        print(f"[入库失败] 导入过程中出现错误：{import_error}")
        return 0
    finally:
        mysql_client.close_db()


if __name__ == "__main__":
    excel_to_mysql()
